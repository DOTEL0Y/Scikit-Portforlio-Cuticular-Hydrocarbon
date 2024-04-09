from openpyxl.workbook import workbook
from openpyxl import load_workbook

wb = load_workbook('ExperimentObservationalData.xlsx')
ws = wb.active

#variable for cages & Dates/days
cage = ws['A2':'A'+str(len('A'))]

print(len(ws['A']))
print(len(ws['B']))
print(len(ws['C']))
print(len(ws['D']))
print(len(ws['F']))
print(len(ws['G']))
print(len(ws['H']))
print(len(ws['I']))
#Create dictionary to contain all information for Cages- Such as
# - How many days of the experiment did it survive? What days? Average?
